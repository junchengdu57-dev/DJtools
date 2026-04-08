#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
完整内存分析器
功能：整合以下所有分析功能
1. Total RAM 分析
2. Proc Meminfo 分析
3. RSS 内存分析（按OOM调整分类）
4. PSS 内存分析（按OOM调整分类）
遍历输入路径，提取所有内存信息并生成综合Excel报告
"""

import os
import re
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 颜色定义 - 类别颜色
CATEGORY_COLORS = {
    'Native': '00FFCCFF',         # 浅蓝色
    'System': '00FFFF99',         # 浅黄色
    'Persistent': '00FF99CC',     # 浅粉色
    'Persistent Service': '00CCFFCC', # 浅绿色
    'Foreground': '00FFCC99',     # 浅橙色
    'Visible': '00CCCCFF',        # 浅紫色
    'Perceptible': '00FFFFCC',    # 浅青色
    'Backup': '00FFE5CC',         # 浅米色
    'Heavy Weight': '00FFCCFF',   # 浅蓝色
    'A Services': '00CCFFFF',     # 浅青蓝
    'Home': '00FF99FF',           # 浅粉紫
    'Previous': '00FFCC99',       # 浅橙色
    'B Services': '00CCFFCC',     # 浅绿色
    'Cached': '00CCCCCC'          # 浅灰色
}

# Total RAM 标签颜色映射（浅色版本，父类仍较深）
TOTAL_RAM_LABEL_COLORS = {
    'Total RAM': '00FFEECC',         # 浅橙色（父类）
    'Free RAM': '00CCFFCC',          # 浅绿色（父类）
    '  cached pss': '00DDFFDD',      # 更浅绿（子类）
    '  cached kernel': '00EEFFEE',   # 最浅绿（子类）
    '  free': '00F5FFF5',            # 超浅绿（子类）
    'DMA-BUF': '00CCCCFF',           # 浅蓝色（父类）
    '  mapped': '00DDDDFF',          # 更浅蓝（子类）
    '  unmapped': '00EEEEFF',        # 最浅蓝（子类）
    'GPU': '00DDCCFF',               # 浅紫色（父类）
    'Used RAM': '00FFCCCC',          # 浅红色（父类）
    '  used pss': '00FFDDDD',        # 更浅红（子类）
    '  kernel': '00FFEEEE',          # 最浅红（子类）
    'Lost RAM': '00FFCCEE',          # 浅紫红色（父类）
    'ZRAM physical used': '00FFFFCC', # 浅黄色（父类）
    '  in swap': '00FFFFDD',         # 更浅黄（子类）
    '  total swap': '00FFFFEE'       # 最浅黄（子类）
}

# debug log开关
debug = False

class CompleteMemoryAnalyzer:
    def __init__(self, input_path, output_file=None):
        self.input_path = input_path
        self.output_file = output_file or f"memory_analysis_complete_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Total RAM 相关数据
        self.round_files = []
        self.overall_memory_data = {}
        
        # Proc Meminfo 相关数据
        self.proc_round_files = []
        self.proc_meminfo_data = {}
        
        # RSS/PSS 相关数据
        self.rss_process_data = {}  # 存储RSS数据
        self.pss_process_data = {}  # 存储PSS数据
        self.category_processes = {}
        self.process_categories_map = {}
        # 使用固定的类别列表
        self.detected_categories = ['Native', 'System', 'Persistent', 'Persistent Service',
                                  'Foreground', 'Visible', 'Perceptible', 'Backup', 'Heavy Weight',
                                  'A Services', 'Home', 'Previous', 'B Services', 'Cached']
        # 为RSS和PSS分别创建类别内存数据结构
        self.rss_category_memory_data = {}  # 存储RSS类别内存数据
        self.pss_category_memory_data = {}  # 存储PSS类别内存数据
    
    def find_dumpsys_files(self):
        """查找所有dumpsys_meminfo格式的文件"""
        dumpsys_files = []
        
        # 递归遍历目录
        for root, dirs, files in os.walk(self.input_path):
            for file in files:
                if 'dumpsys_meminfo_round' in file or 'dumpsys_meminfo.txt' in file:
                    file_path = os.path.join(root, file)
                    dumpsys_files.append(file_path)
                    if debug:
                        print("递归遍历至: " + file_path)
        
        # 按文件名排序，确保轮次顺序正确
        dumpsys_files.sort(key=lambda p: int(re.search(r'round_(\d+)', os.path.basename(os.path.dirname(p))).group(1)) if re.search(r'round_(\d+)', os.path.basename(os.path.dirname(p))) else float('inf'))
        self.round_files = dumpsys_files
        if debug:
            print(self.round_files)
        return dumpsys_files
    
    def find_proc_meminfo_files(self):
        """查找所有 proc_meminfo.txt 文件并按轮次排序"""
        files = []
        for root, _, fs in os.walk(self.input_path):
            for fn in fs:
                if 'proc_meminfo' in fn:
                    files.append(os.path.join(root, fn))
        def round_key(p):
            m = re.search(r'round_(\d+)', os.path.basename(os.path.dirname(p)))
            return int(m.group(1)) if m else float('inf')
        files.sort(key=round_key)
        self.proc_round_files = files
        return files
    
    def _parse_overall_memory(self, lines):
        """解析 dumpsys meminfo 中的总体内存信息"""
        def to_int(s):
            s = s.replace(',', '').strip(); return int(s) if s.isdigit() else 0
        mem = {}
        for raw in lines:
            line = raw.strip()
            m = re.match(r'^Total RAM:\s*([\d,]+)K', line)
            if m: mem['Total RAM'] = to_int(m.group(1)); continue
            m = re.match(r'^\s*Free RAM:\s*([\d,]+)K\s*\(\s*([\d,]+)K cached pss \+\s*([\d,]+)K cached kernel \+\s*([\d,]+)K free\)', line)
            if m:
                mem['Free RAM'] = {'total': to_int(m.group(1)), 'cached pss': to_int(m.group(2)), 'cached kernel': to_int(m.group(3)), 'free': to_int(m.group(4))}; continue
            m = re.match(r'^\s*DMA-BUF:\s*([\d,]+)K\s*\(\s*([\d,]+)K mapped \+\s*([\d,]+)K unmapped\)', line)
            if m:
                mem['DMA-BUF'] = {'total': to_int(m.group(1)), 'mapped': to_int(m.group(2)), 'unmapped': to_int(m.group(3))}; continue
            m = re.match(r'^\s*GPU:\s*([\d,]+)K\s*\(\s*([\d,]+)K dmabuf \+\s*([\d,]+)K private\)', line)
            if m:
                mem['GPU'] = {'total': to_int(m.group(1)), 'dmabuf': to_int(m.group(2)), 'private': to_int(m.group(3))}; continue
            m = re.match(r'^\s*Used RAM:\s*([\d,]+)K\s*\(\s*([\d,]+)K used pss \+\s*([\d,]+)K kernel\)', line)
            if m:
                mem['Used RAM'] = {'total': to_int(m.group(1)), 'used pss': to_int(m.group(2)), 'kernel': to_int(m.group(3))}; continue
            m = re.match(r'^\s*Lost RAM:\s*([\d,]+)K', line)
            if m: mem['Lost RAM'] = to_int(m.group(1)); continue
            m = re.match(r'^\s*ZRAM:\s*([\d,]+)K physical used for\s*([\d,]+)K in swap\s*\(\s*([\d,]+)K total swap\)', line)
            if m:
                mem['ZRAM'] = {'physical used': to_int(m.group(1)), 'in swap': to_int(m.group(2)), 'total swap': to_int(m.group(3))}; continue
        return mem
    
    def extract_overall_memory(self, file_path):
        """从文件中提取总体内存信息"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                lines = f.read().split('\n')
            return self._parse_overall_memory(lines)
        except Exception as e:
            print(f"读取文件 {file_path} 时出错: {e}")
            return {}
    
    def _parse_proc_meminfo_file(self, file_path):
        """解析单个 proc_meminfo.txt 的关键字段，返回字典（单位kB，整数）"""
        target_keys = {
            'MemTotal','MemFree','MemAvailable','Buffers','Cached','SwapCached','Active','Inactive',
            'Active(anon)','Inactive(anon)','Active(file)','Inactive(file)','Unevictable','Mlocked',
            'SwapTotal','SwapFree','Dirty','Writeback','AnonPages','Mapped','Shmem','KReclaimable',
            'Slab','SReclaimable','SUnreclaim','KernelStack','ShadowCallStack','PageTables','SecPageTables',
            'NFS_Unstable','Bounce','WritebackTmp','CommitLimit','Committed_AS','VmallocTotal','VmallocUsed',
            'VmallocChunk','Percpu','AnonHugePages','ShmemHugePages','ShmemPmdMapped','FileHugePages',
            'FilePmdMapped','CmaTotal','CmaFree'
        }
        data = {}
        try:
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                for raw in f:
                    if ':' not in raw:
                        continue
                    key = raw.split(':', 1)[0].strip()
                    if key not in target_keys:
                        continue
                    m = re.search(r'([0-9,]+)\s*kB', raw)
                    if m:
                        val = int(m.group(1).replace(',', ''))
                        data[key] = val
        except Exception as e:
            print(f"读取文件 {file_path} 时出错: {e}")
        return data
    
    def extract_memory_data(self, file_path):
        """从单个文件中提取内存数据（从OOM adjustment部分提取）"""
        rss_memory = {}
        pss_memory = {}
        process_categories = {}
        rss_category_memory = {}
        pss_category_memory = {}
        
        try:
            # 尝试读取文件内容
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
            
            lines = content.split('\n')
            
            # 1. 提取RSS数据
            in_rss_oom_section = False
            current_category = None
            
            for line in lines:
                line = line.strip()
                
                if line.startswith('Total RSS by OOM adjustment:'):
                    in_rss_oom_section = True
                    continue
                
                if in_rss_oom_section and line.startswith('Total RSS by category:'):
                    in_rss_oom_section = False
                    break
                
                if in_rss_oom_section:
                    # 类别行处理
                    if 'K:' in line and '(pid ' not in line and line:
                        parts = line.split('K:')
                        if len(parts) > 1:
                            memory_str = parts[0].strip().replace(',', '')
                            if memory_str.isdigit():
                                category_memory_value = int(memory_str)
                                # 兼容类别后缀形如 "( xxxK in swap)"，并规范别名
                                raw_name = parts[1].strip()
                                clean_name = re.sub(r"\s*\(.*\)\s*$", "", raw_name).strip()
                                if clean_name.lower() in ('perceptible low', 'perceptible medium'):
                                    clean_name = 'Perceptible'
                                current_category = clean_name
                                rss_category_memory[clean_name] = rss_category_memory.get(clean_name, 0) + category_memory_value
                    # 进程行处理
                    elif 'K:' in line and '(pid ' in line:
                        parts = line.split('K:')
                        if len(parts) > 1:
                            memory_str = parts[0].strip().replace(',', '')
                            if memory_str.isdigit():
                                memory_value = int(memory_str)
                                process_part = parts[1].strip()
                                process_match = re.match(r'([\w\._-]+(?:[:\w]+)?)', process_part)
                                if process_match and current_category:
                                    process = process_match.group(1)
                                    rss_memory[process] = memory_value
                                    process_categories[process] = current_category
            
            # 2. 提取PSS数据
            in_pss_oom_section = False
            current_category = None
            
            for line in lines:
                line = line.strip()
                
                if line.startswith('Total PSS by OOM adjustment:'):
                    in_pss_oom_section = True
                    continue
                
                if in_pss_oom_section and line.startswith('Total PSS by category:'):
                    in_pss_oom_section = False
                    break
                
                if in_pss_oom_section:
                    # 类别行处理
                    if 'K:' in line and '(pid ' not in line and line:
                        parts = line.split('K:')
                        if len(parts) > 1:
                            memory_str = parts[0].strip().replace(',', '')
                            if memory_str.isdigit():
                                category_memory_value = int(memory_str)
                                # 兼容类别后缀形如 "( xxxK in swap)"，并规范别名
                                raw_name = parts[1].strip()
                                clean_name = re.sub(r"\s*\(.*\)\s*$", "", raw_name).strip()
                                if clean_name.lower() in ('perceptible low', 'perceptible medium'):
                                    clean_name = 'Perceptible'
                                current_category = clean_name
                                pss_category_memory[clean_name] = pss_category_memory.get(clean_name, 0) + category_memory_value
                    # 进程行处理
                    elif 'K:' in line and '(pid ' in line:
                        parts = line.split('K:')
                        if len(parts) > 1:
                            memory_str = parts[0].strip().replace(',', '')
                            if memory_str.isdigit():
                                memory_value = int(memory_str)
                                process_part = parts[1].strip()
                                process_match = re.match(r'([\w\._-]+(?:[:\w]+)?)', process_part)
                                if process_match and current_category:
                                    process = process_match.group(1)
                                    pss_memory[process] = memory_value
                                    if process not in process_categories:
                                        process_categories[process] = current_category
        
        except Exception as e:
            print(f"读取文件 {file_path} 时出错: {e}")
        
        return rss_memory, pss_memory, process_categories, rss_category_memory, pss_category_memory
    
    def classify_process(self, process_name):
        """简化的进程分类方法"""
        # 检查是否已经分类过
        if process_name in self.process_categories_map:
            return self.process_categories_map[process_name]
        
        name_lower = process_name.lower()
        
        # 基于进程名的规则分类
        # Native进程
        if any(native_keyword in process_name for native_keyword in ['/system/bin/', '/vendor/bin/', 'logd', 'surfaceflinger', 'zygote']):
            return self._assign_category(process_name, 'Native')
        
        # Persistent Service
        if 'service' in name_lower and any(persistent_keyword in name_lower for persistent_keyword in ['persistent', 'system']):
            return self._assign_category(process_name, 'Persistent Service')
        
        # A Services：前台或活跃服务
        if 'service' in name_lower and not any(bg_kw in name_lower for bg_kw in ['background', 'bg.', 'b.']) and 'persistent' not in name_lower:
            return self._assign_category(process_name, 'A Services')
        
        # 前台应用
        if any(foreground_keyword in process_name for foreground_keyword in ['com.instagram', 'com.facebook', 'com.tencent', 'com.baidu']):
            return self._assign_category(process_name, 'Foreground')
        
        # Home 桌面/启动器
        if any(home_keyword in name_lower for home_keyword in ['launcher', 'nexuslauncher', 'com.android.launcher', 'home']):
            return self._assign_category(process_name, 'Home')
        
        # 后台服务
        if any(service_keyword in process_name for service_keyword in ['background', 'bg.', 'b.']):
            return self._assign_category(process_name, 'B Services')
        
        # Backup 备份相关
        if 'backup' in name_lower or 'com.android.backup' in name_lower:
            return self._assign_category(process_name, 'Backup')
        
        # System进程
        if ('system' in name_lower or 
            any(android_keyword in process_name for android_keyword in ['android.process.', 'com.android.', 'android.'])):
            return self._assign_category(process_name, 'System')
        
        # Persistent进程
        if 'persistent' in name_lower:
            return self._assign_category(process_name, 'Persistent')
        
        # 其他类别简化处理
        if 'heavy' in name_lower or 'weight' in name_lower:
            return self._assign_category(process_name, 'Heavy Weight')
        
        if 'visible' in name_lower or 'ui' in name_lower:
            return self._assign_category(process_name, 'Visible')
        
        if 'perceptible' in name_lower:
            return self._assign_category(process_name, 'Perceptible')
        
        if 'previous' in name_lower:
            return self._assign_category(process_name, 'Previous')
        
        # 默认分类为Cached
        return self._assign_category(process_name, 'Cached')
    
    def _assign_category(self, process_name, category):
        """简化的进程分类分配"""
        # 存储分类映射
        self.process_categories_map[process_name] = category
        
        # 确保类别在category_processes中存在
        if category not in self.category_processes:
            self.category_processes[category] = []
        
        # 只添加不存在的进程
        if process_name not in self.category_processes[category]:
            self.category_processes[category].append(process_name)
        
        return category
    
    def generate_total_ram_sheet(self, wb):
        """生成 Total RAM 数据工作表"""
        title = 'Total RAM'
        for sheet in wb.sheetnames:
            if sheet == title:
                del wb[sheet]; break
        ws = wb.create_sheet(title=title)
        ws.column_dimensions['A'].width = 22.25  # 设置A列列宽为22.25
        row = 1
        ws.cell(row=row, column=1).value = '系统内存标签占用'
        ws.cell(row=row, column=1).font = Font(bold=True, size=14, color='FFFFFF')  # 白色字体以确保可读性
        ws.cell(row=row, column=1).fill = PatternFill(start_color='00483D8B', end_color='00483D8B', fill_type='solid')  # 暗板岩蓝背景
        row += 1
        ws.cell(row=row, column=1).value = '标签'
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = '平均值(KB)'
        ws.cell(row=row, column=2).font = Font(bold=True, color='FF0000')
        ws.cell(row=row, column=2).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for i, file_path in enumerate(self.round_files):
            ws.cell(row=row, column=i+3).value = f"轮次 {i+1} ({os.path.basename(file_path)})"
            ws.cell(row=row, column=i+3).font = Font(bold=True)
        row += 1
        labels = [
            'Total RAM',
            'Free RAM',
            '  cached pss',
            '  cached kernel',
            '  free',
            'DMA-BUF',
            '  mapped',
            '  unmapped',
            'GPU',
            '  dmabuf',
            '  private',
            'Used RAM',
            '  used pss',
            '  kernel',
            'Lost RAM',
            'ZRAM physical used',
            '  in swap',
            '  total swap'
        ]
        for label in labels:
            values = []
            for round_idx in range(len(self.round_files)):
                d = self.overall_memory_data.get(round_idx, {})
                v = 0
                if label == 'Total RAM':
                    v = d.get('Total RAM', 0)
                elif label == 'Free RAM':
                    v = d.get('Free RAM', {}).get('total', 0)
                elif label.strip() in ['cached pss', 'cached kernel', 'free']:
                    v = d.get('Free RAM', {}).get(label.strip(), 0)
                elif label == 'DMA-BUF':
                    v = d.get('DMA-BUF', {}).get('total', 0)
                elif label.strip() in ['mapped', 'unmapped']:
                    v = d.get('DMA-BUF', {}).get(label.strip(), 0)
                elif label == 'GPU':
                    v = d.get('GPU', {}).get('total', 0)
                elif label.strip() in ['dmabuf', 'private']:
                    v = d.get('GPU', {}).get(label.strip(), 0)
                elif label == 'Used RAM':
                    v = d.get('Used RAM', {}).get('total', 0)
                elif label.strip() in ['used pss', 'kernel']:
                    v = d.get('Used RAM', {}).get(label.strip(), 0)
                elif label == 'Lost RAM':
                    v = d.get('Lost RAM', 0)
                elif label == 'ZRAM physical used':
                    v = d.get('ZRAM', {}).get('physical used', 0)
                elif label.strip() in ['in swap', 'total swap']:
                    v = d.get('ZRAM', {}).get(label.strip(), 0)
                values.append(v or 0)
            valid = [vv for vv in values if vv > 0]
            avg_value = round(sum(valid)/len(valid)) if valid else 0
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=2).value = avg_value
            ws.cell(row=row, column=2).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            ws.cell(row=row, column=2).font = Font(color='FF0000')
            # 为标签列添加颜色
            if label in TOTAL_RAM_LABEL_COLORS:
                ws.cell(row=row, column=1).fill = PatternFill(start_color=TOTAL_RAM_LABEL_COLORS[label], end_color=TOTAL_RAM_LABEL_COLORS[label], fill_type='solid')
                ws.cell(row=row, column=1).font = Font(bold=True, color='000000')  # 黑色字体，提高可读性
            for i in range(len(self.round_files)):
                ws.cell(row=row, column=i+3).value = values[i]
            row += 1
    
    def generate_proc_meminfo_sheet(self, wb):
        """生成 Proc Meminfo 数据工作表"""
        title = 'Proc Meminfo'
        for sheet in wb.sheetnames:
            if sheet == title:
                del wb[sheet]; break
        ws = wb.create_sheet(title=title)
        ws.column_dimensions['A'].width = 22.25  # 设置A列列宽为22.25
        row = 1
        ws.cell(row=row, column=1).value = 'Proc Meminfo'
        ws.cell(row=row, column=1).font = Font(bold=True, size=14, color='FFFFFF')  # 白色字体以确保可读性
        ws.cell(row=row, column=1).fill = PatternFill(start_color='00483D8B', end_color='00483D8B', fill_type='solid')  # 暗板岩蓝背景
        row += 1
        # 表头
        ws.cell(row=row, column=1).value = '标签'
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = '平均值(kB)'
        ws.cell(row=row, column=2).font = Font(bold=True, color='FF0000')
        ws.cell(row=row, column=2).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for i, fp in enumerate(self.proc_round_files):
            ws.cell(row=row, column=i+3).value = f"轮次 {i+1} ({os.path.basename(fp)})"
            ws.cell(row=row, column=i+3).font = Font(bold=True)
        row += 1
        labels = [
            'MemTotal','MemFree','MemAvailable','Buffers','Cached','SwapCached','Active','Inactive',
            'Active(anon)','Inactive(anon)','Active(file)','Inactive(file)','Unevictable','Mlocked',
            'SwapTotal','SwapFree','Dirty','Writeback','AnonPages','Mapped','Shmem','KReclaimable',
            'Slab','SReclaimable','SUnreclaim','KernelStack','ShadowCallStack','PageTables','SecPageTables',
            'NFS_Unstable','Bounce','WritebackTmp','CommitLimit','Committed_AS','VmallocTotal','VmallocUsed',
            'VmallocChunk','Percpu','AnonHugePages','ShmemHugePages','ShmemPmdMapped','FileHugePages',
            'FilePmdMapped','CmaTotal','CmaFree'
        ]
        rounds = len(self.proc_round_files)
        for label in labels:
            values = []
            for idx in range(rounds):
                values.append(self.proc_meminfo_data.get(idx, {}).get(label, 0))
            valid = [v for v in values if v > 0]
            avg_value = round(sum(valid)/len(valid)) if valid else 0
            ws.cell(row=row, column=1).value = label
            # 为Proc Meminfo的标签列添加浅灰色背景，提高可读性
            ws.cell(row=row, column=1).fill = PatternFill(start_color='00F0F0F0', end_color='00F0F0F0', fill_type='solid')
            ws.cell(row=row, column=1).font = Font(bold=True, color='000000')  # 黑色粗体字体
            ws.cell(row=row, column=2).value = avg_value
            ws.cell(row=row, column=2).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            ws.cell(row=row, column=2).font = Font(color='FF0000')
            for i in range(rounds):
                ws.cell(row=row, column=i+3).value = values[i]
            row += 1
    
    def _generate_sheet(self, wb, title, process_data):
        """简化的工作表生成方法（用于RSS和PSS分析）"""
        # 移除已存在的工作表
        for sheet in wb.sheetnames:
            if sheet == title:
                del wb[sheet]
                break
                
        ws = wb.create_sheet(title=title)
        row = 1
        
        # 添加类别内存占用标题
        ws.cell(row=row, column=1).value = f'{title.split(" ")[0]} 类别内存占用（KB）'
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 1
        
        # 表头设置
        ws.cell(row=row, column=1).value = '类别'
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = '平均值(KB)'
        ws.cell(row=row, column=2).font = Font(bold=True, color='FF0000')
        ws.cell(row=row, column=2).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid")
        
        # 设置轮次标题
        for i, file_path in enumerate(self.round_files):
            file_name = os.path.basename(file_path)
            ws.cell(row=row, column=i + 3).value = f'轮次 {i+1} ({file_name})'
            ws.cell(row=row, column=i + 3).font = Font(bold=True)
        
        row += 1
        
        # 选择类别内存数据
        is_rss_sheet = title.startswith('RSS')
        category_memory_data = self.rss_category_memory_data if is_rss_sheet else self.pss_category_memory_data
        
        # 写入类别内存数据
        for category in self.detected_categories:
            # 计算类别值和平均值
            category_values = []
            for round_idx in range(len(self.round_files)):
                value = category_memory_data.get(round_idx, {}).get(category, 0)
                category_values.append(value)
            
            valid_values = [v for v in category_values if v > 0]
            avg_value = round(sum(valid_values) / len(valid_values)) if valid_values else 0
            
            # 写入数据
            ws.cell(row=row, column=1).value = category
            ws.cell(row=row, column=2).value = avg_value
            ws.cell(row=row, column=2).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid")
            ws.cell(row=row, column=2).font = Font(color='FF0000')
            
            for i in range(len(self.round_files)):
                ws.cell(row=row, column=i + 3).value = category_values[i]
            
            row += 1
        
        # 空行分隔
        row += 2
        
        # 进程数据标题
        ws.cell(row=row, column=1).value = '进程类别/名称'
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = '平均值(KB)'
        ws.cell(row=row, column=2).font = Font(bold=True)
        
        for i, file_path in enumerate(self.round_files):
            ws.cell(row=row, column=i + 3).value = f'轮次 {i+1}'
            ws.cell(row=row, column=i + 3).font = Font(bold=True)
        
        row += 1
        
        # 写入进程数据
        total_process_count = 0
        
        for category in self.detected_categories:
            process_list = self.category_processes.get(category, [])
            if process_list:
                total_process_count += len(process_list)
                # 类别标题
                ws.cell(row=row, column=1).value = f"{category} (共{len(process_list)}个进程)"
                ws.cell(row=row, column=1).font = Font(bold=True)
                
                # 设置颜色
                if category in CATEGORY_COLORS:
                    fill = PatternFill(start_color=CATEGORY_COLORS[category], end_color=CATEGORY_COLORS[category], fill_type="solid")
                    ws.cell(row=row, column=1).fill = fill
                
                row += 1
                
                # 计算并排序进程数据
                process_with_avg = []
                for process in process_list:
                    values = []
                    for i in range(len(self.round_files)):
                        value = process_data.get(category, {}).get(process, {}).get(i, 0)
                        if value > 0:
                            values.append(value)
                    
                    avg_value = round(sum(values) / len(values)) if values else 0
                    process_with_avg.append((process, avg_value))
                
                # 按平均值排序
                process_with_avg.sort(key=lambda x: x[1], reverse=True)
                
                # 写入进程数据
                for process, avg_value in process_with_avg:
                    ws.cell(row=row, column=1).value = process
                    ws.cell(row=row, column=2).value = avg_value
                    ws.cell(row=row, column=2).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid")
                    ws.cell(row=row, column=2).font = Font(color='FF0000')
                    
                    for i in range(len(self.round_files)):
                        memory_value = process_data.get(category, {}).get(process, {}).get(i, 0)
                        ws.cell(row=row, column=i + 3).value = memory_value
                    
                    row += 1
                
                row += 1  # 类别间隔
        
        # 汇总信息
        ws.cell(row=row, column=1).value = "汇总信息"
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1).value = f"总文件数: {len(self.round_files)}"
        ws.cell(row=row, column=2).value = f"总进程数: {total_process_count}"
        
        # 调整列宽
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 12
        for i in range(len(self.round_files)):
            col_idx = i + 3
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15
        
        return total_process_count
    
    def generate_excel(self):
        """生成包含所有分析功能的Excel报告"""
        try:
            # 创建新的工作簿
            wb = Workbook()
            # 删除默认工作表
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            
            # 生成 Total RAM 工作表
            print("生成 Total RAM 工作表...")
            self.generate_total_ram_sheet(wb)
            
            # 生成 Proc Meminfo 工作表
            print("生成 Proc Meminfo 工作表...")
            self.generate_proc_meminfo_sheet(wb)
            
            # 生成 RSS 内存分析工作表
            print("生成 RSS 内存分析工作表...")
            rss_process_count = self._generate_sheet(wb, "RSS内存分析", self.rss_process_data)
            
            # 生成 PSS 内存分析工作表
            print("生成 PSS 内存分析工作表...")
            pss_process_count = self._generate_sheet(wb, "PSS内存分析", self.pss_process_data)
            
            # 确保输出目录存在
            output_dir = os.path.dirname(self.output_file)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # 保存文件
            wb.save(self.output_file)
            print(f"Excel文件已保存到: {self.output_file}")
            print(f"RSS分析了 {rss_process_count} 个进程")
            print(f"PSS分析了 {pss_process_count} 个进程")
            return True
        except Exception as e:
            print(f"保存Excel文件时出错: {e}")
            return False
    
    def analyze_all_files(self, directory):
        """分析所有文件并生成完整的内存报告"""
        print("完整内存分析器 开始")
        
        self.input_path = directory
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.output_file = f"memory_analysis_complete_{timestamp}.xlsx"
        
        # 初始化数据结构
        self.rss_process_data = {}
        self.pss_process_data = {}
        self.process_categories_map = {}
        self.category_processes = {}
        self.round_files = []
        self.rss_category_memory_data = {}
        self.pss_category_memory_data = {}
        self.overall_memory_data = {}
        self.proc_meminfo_data = {}
        
        # 查找dumpsys文件
        dumpsys_files = self.find_dumpsys_files()
        if not dumpsys_files:
            print("未找到dumpsys_meminfo文件")
        else:
            print(f"找到 {len(dumpsys_files)} 个dumpsys_meminfo文件")
            
            # 处理每个文件，提取所有需要的数据
            for i, file_path in enumerate(dumpsys_files):
                print(f"正在处理文件 {i+1}/{len(dumpsys_files)}: {os.path.basename(file_path)}")
                
                # 提取总体内存信息 (Total RAM)
                overall = self.extract_overall_memory(file_path)
                self.overall_memory_data[i] = overall.copy() if overall else {}
                
                # 提取RSS/PSS数据
                process_rss, process_pss, process_categories, rss_category_memory, pss_category_memory = self.extract_memory_data(file_path)
                
                # 更新进程分类映射
                self.process_categories_map.update(process_categories)
                
                # 保存类别内存数据
                self.rss_category_memory_data[i] = rss_category_memory.copy()
                self.pss_category_memory_data[i] = pss_category_memory.copy()
                
                # 处理RSS数据
                for process, rss in process_rss.items():
                    if process in process_categories:
                        category = process_categories[process]
                        # 确保类别数据结构
                        if category not in self.category_processes:
                            self.category_processes[category] = []
                        if process not in self.category_processes[category]:
                            self.category_processes[category].append(process)
                    else:
                        # 使用简化的分类方法
                        category = self.classify_process(process)
                    
                    # 保存数据并填充历史数据
                    if category not in self.rss_process_data:
                        self.rss_process_data[category] = {}
                    if process not in self.rss_process_data[category]:
                        self.rss_process_data[category][process] = {}
                    
                    self.rss_process_data[category][process][i] = rss
                    for prev_round in range(i):
                        if prev_round not in self.rss_process_data[category][process]:
                            self.rss_process_data[category][process][prev_round] = 0
                
                # 处理PSS数据
                for process, pss in process_pss.items():
                    if process in process_categories:
                        category = process_categories[process]
                        if category not in self.category_processes:
                            self.category_processes[category] = []
                        if process not in self.category_processes[category]:
                            self.category_processes[category].append(process)
                    else:
                        # 使用已有的分类或重新分类
                        category = self.process_categories_map.get(process, self.classify_process(process))
                    
                    # 保存数据并填充历史数据
                    if category not in self.pss_process_data:
                        self.pss_process_data[category] = {}
                    if process not in self.pss_process_data[category]:
                        self.pss_process_data[category][process] = {}
                    
                    self.pss_process_data[category][process][i] = pss
                    for prev_round in range(i):
                        if prev_round not in self.pss_process_data[category][process]:
                            self.pss_process_data[category][process][prev_round] = 0
                
                # 填充缺失的0值
                for category in list(self.rss_process_data.keys()):
                    for process in list(self.rss_process_data[category].keys()):
                        if i not in self.rss_process_data[category][process]:
                            self.rss_process_data[category][process][i] = 0
                
                for category in list(self.pss_process_data.keys()):
                    for process in list(self.pss_process_data[category].keys()):
                        if i not in self.pss_process_data[category][process]:
                            self.pss_process_data[category][process][i] = 0
        
        # 处理proc_meminfo文件
        proc_files = self.find_proc_meminfo_files()
        if proc_files:
            print(f"找到 {len(proc_files)} 个proc_meminfo文件")
            for j, pf in enumerate(proc_files):
                print(f"正在处理proc文件 {j+1}/{len(proc_files)}: {os.path.basename(pf)}")
                metrics = self._parse_proc_meminfo_file(pf)
                self.proc_meminfo_data[j] = metrics
        else:
            print('未找到proc_meminfo文件')
        
        # 生成Excel报告
        self.generate_excel()
        
        print("完整内存分析器 完成")
        return True

def main():
    """主函数"""
    # 获取输入参数
    input_path = sys.argv[1] if len(sys.argv) > 1 else os.getcwd()
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    # 运行分析器
    analyzer = CompleteMemoryAnalyzer(input_path, output_file)
    analyzer.analyze_all_files(input_path)

if __name__ == "__main__":
    main()